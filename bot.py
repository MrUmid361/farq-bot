"""
Buxgalteriya yordamchi Telegram bot.

Bo'limlar:
  - Fayllarni solishtirish (mavjud)
  - Excel faylni tekshirish (arifmetik xatolarni topish)
  - Foyda solig'i hisob-kitobi
  - QQS kalkulyatori
  - Ish haqi kalkulyatori
  - Penya (jarima) kalkulyatori
  - Amortizatsiya kalkulyatori
  - CBU rasmiy valyuta kursi
  - Hisob-faktura generator
  - Ish kunlari kalkulyatori
  - Kredit/lizing to'lov jadvali
  - Moliyaviy koeffitsientlar
  - YATT uchun soliq (ma'lumot)
  - Fakturalar asosida chorak bo'yicha foyda solig'i hisoboti

O'RNATISH:
    pip install python-telegram-bot==21.* pandas openpyxl beautifulsoup4 lxml xlrd requests

ISHGA TUSHIRISH:
    export BOT_TOKEN="123456:ABC-tokeningiz"
    python bot.py
"""

import io
import os
import logging
import re
from dataclasses import dataclass, field
from datetime import datetime

import pandas as pd
import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from telegram import Update, Document, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    Application,
    CommandHandler,
    MessageHandler,
    CallbackQueryHandler,
    ContextTypes,
    ConversationHandler,
    filters,
)

logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    level=logging.INFO,
)
logger = logging.getLogger(__name__)

BOT_TOKEN = os.environ.get("BOT_TOKEN", "")

# ---------------- Conversation holatlari ----------------
MENU, TEXT_INPUT, WAITING_FILE_1, WAITING_FILE_2, WAITING_VALIDATE_FILE, WAITING_TAX_REPORT_FILE = range(6)


# ================================================================
#                 1) FAYLLARNI SOLISHTIRISH (mavjud)
# ================================================================

NAME_KEYWORDS = ["номенклатура", "маҳсулот", "название", "наименование", "nomi", "товар"]
CODE_KEYWORDS = ["код", "kod"]
QTY_KEYWORDS = ["количество", "миқдор", "кол-во", "miqdor"]
PRICE_KEYWORDS = ["цена", "нарх", "нарҳ", "narx"]
SUM_KEYWORDS = ["сумма", "қиймат", "summa", "qiymat"]
EXCLUDE_IN_SUM = ["ндс", "ққс", "жами", "vat", "qqsni"]


@dataclass
class ParsedFile:
    df: pd.DataFrame
    raw_columns: list = field(default_factory=list)


def _norm(s):
    return str(s).strip().lower()


def _to_number(x):
    if x is None:
        return None
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).strip().replace("\xa0", "").replace(" ", "")
    s = s.replace(",", ".")
    s = re.sub(r"[^0-9.\-]", "", s)
    if s in ("", "-", "."):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _pick_column(columns, keywords, exclude=None):
    exclude = exclude or []
    for col in columns:
        c = _norm(col)
        if any(k in c for k in keywords) and not any(e in c for e in exclude):
            return col
    return None


def read_xlsx(path: str) -> ParsedFile:
    df_raw = pd.read_excel(path)
    cols = list(df_raw.columns)

    name_col = _pick_column(cols, NAME_KEYWORDS)
    code_col = _pick_column(cols, CODE_KEYWORDS)
    qty_col = _pick_column(cols, QTY_KEYWORDS)
    price_col = _pick_column(cols, PRICE_KEYWORDS)
    sum_col = _pick_column(cols, SUM_KEYWORDS, exclude=EXCLUDE_IN_SUM)

    if not all([name_col, qty_col, price_col, sum_col]):
        raise ValueError(
            "Ustunlarni avtomatik aniqlab bo'lmadi. Topilgan ustunlar: " + ", ".join(map(str, cols))
        )

    out = pd.DataFrame()
    out["name"] = df_raw[name_col].astype(str).str.strip()
    out["code"] = df_raw[code_col].astype(str).str.strip().str.lstrip("0") if code_col else ""
    out["qty"] = df_raw[qty_col].apply(_to_number)
    out["price"] = df_raw[price_col].apply(_to_number)
    out["sum"] = df_raw[sum_col].apply(_to_number)
    out = out.dropna(subset=["qty", "price", "sum"])
    return ParsedFile(df=out, raw_columns=cols)


def _find_product_table(soup: BeautifulSoup):
    best = None
    best_rows = 0
    for table in soup.find_all("table"):
        rows = table.find_all("tr")
        numeric_rows = 0
        for r in rows:
            cells = r.find_all(["td", "th"])
            if cells and cells[0].get_text(strip=True).isdigit():
                numeric_rows += 1
        if numeric_rows > best_rows:
            best_rows = numeric_rows
            best = table
    return best, best_rows


def read_html_xls(path: str) -> ParsedFile:
    with open(path, "rb") as f:
        raw = f.read()

    encoding = "utf-8"
    m = re.search(rb"charset=([\w-]+)", raw[:2000], re.IGNORECASE)
    if m:
        encoding = m.group(1).decode("ascii", errors="ignore")

    html = raw.decode(encoding, errors="ignore")
    soup = BeautifulSoup(html, "lxml")

    table, n_rows = _find_product_table(soup)
    if table is None or n_rows < 2:
        raise ValueError("Faylda mahsulotlar jadvali topilmadi.")

    data_rows = []
    for r in table.find_all("tr"):
        cells = [c.get_text(strip=True) for c in r.find_all(["td", "th"])]
        if cells and cells[0].isdigit() and len(cells) >= 7:
            data_rows.append(cells)

    if not data_rows:
        raise ValueError("Faylda mahsulot qatorlari topilmadi.")

    ncols = len(data_rows[0])
    name_idx = 1
    code_idx = 2 if ncols > 2 else None
    qty_idx = 4 if ncols > 4 else 3
    price_idx = 5 if ncols > 5 else 4
    sum_idx = 6 if ncols > 6 else 5

    records = []
    for cells in data_rows:
        try:
            name = cells[name_idx]
            code = ""
            if code_idx is not None and code_idx < len(cells):
                code = cells[code_idx].split("-")[0].strip().lstrip("0")
            qty = _to_number(cells[qty_idx])
            price = _to_number(cells[price_idx])
            total = _to_number(cells[sum_idx])
            if qty is None or price is None or total is None:
                continue
            records.append((name, code, qty, price, total))
        except IndexError:
            continue

    df = pd.DataFrame(records, columns=["name", "code", "qty", "price", "sum"])
    return ParsedFile(df=df, raw_columns=[f"col{i}" for i in range(ncols)])


def read_any(path: str) -> ParsedFile:
    lower = path.lower()
    if lower.endswith(".xlsx") or lower.endswith(".xlsm"):
        return read_xlsx(path)

    if lower.endswith(".xls"):
        try:
            return read_xlsx(path)
        except Exception:
            return read_html_xls(path)

    raise ValueError("Faqat .xlsx yoki .xls fayllar qo'llab-quvvatlanadi.")


def compare_files(pf1: ParsedFile, pf2: ParsedFile) -> tuple[bytes, dict]:
    from collections import Counter

    df1, df2 = pf1.df.copy(), pf2.df.copy()

    def key(row):
        return (round(row["qty"], 3), round(row["price"], 2))

    c1 = Counter(key(r) for _, r in df1.iterrows())
    c2 = Counter(key(r) for _, r in df2.iterrows())

    only1 = c1 - c2
    only2 = c2 - c1

    rows_out = []
    for k, v in only1.items():
        matches = df1[(df1["qty"].round(3) == k[0]) & (df1["price"].round(2) == k[1])]
        for _, m in matches.iterrows():
            rows_out.append(("1-faylda bor, 2-faylda yo'q", m["name"], m["code"], m["qty"], m["price"], m["sum"]))
    for k, v in only2.items():
        matches = df2[(df2["qty"].round(3) == k[0]) & (df2["price"].round(2) == k[1])]
        for _, m in matches.iterrows():
            rows_out.append(("2-faylda bor, 1-faylda yo'q", m["name"], m["code"], m["qty"], m["price"], m["sum"]))

    diff_df = pd.DataFrame(rows_out, columns=["Manba", "Nomi", "Kod", "Miqdor", "Narx", "Summa"])
    diff_df = diff_df.sort_values(["Manba", "Nomi"]).reset_index(drop=True)

    total1 = df1["sum"].sum()
    total2 = df2["sum"].sum()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Xulosa"
    bold = Font(name="Arial", bold=True)
    title_font = Font(name="Arial", bold=True, size=14)

    ws["A1"] = "Fayllar orasidagi farq hisoboti"
    ws["A1"].font = title_font
    ws.merge_cells("A1:D1")

    ws["A3"] = "1-fayl qatorlar soni"
    ws["B3"] = len(df1)
    ws["A4"] = "2-fayl qatorlar soni"
    ws["B4"] = len(df2)
    ws["A5"] = "1-fayl jami summasi"
    ws["B5"] = round(total1, 2)
    ws["A6"] = "2-fayl jami summasi"
    ws["B6"] = round(total2, 2)
    ws["A7"] = "Umumiy farq (2-fayl − 1-fayl)"
    ws["B7"] = "=B6-B5"
    for r in (3, 4, 5, 6, 7):
        ws[f"A{r}"].font = Font(name="Arial")
    ws["A7"].font = bold
    ws["B7"].font = bold
    for r in (5, 6, 7):
        ws[f"B{r}"].number_format = "#,##0.00"
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 18

    ws2 = wb.create_sheet("Farqlar")
    headers = ["№", "Manba", "Nomi", "Kod", "Miqdor", "Narx", "Summa"]
    ws2.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws2.cell(row=1, column=c)
        cell.font = bold
        cell.fill = PatternFill("solid", fgColor="D9E1F2")

    for i, row in diff_df.iterrows():
        r = i + 2
        ws2.cell(row=r, column=1, value=i + 1)
        ws2.cell(row=r, column=2, value=row["Manba"])
        ws2.cell(row=r, column=3, value=row["Nomi"])
        ws2.cell(row=r, column=4, value=row["Kod"])
        ws2.cell(row=r, column=5, value=row["Miqdor"])
        ws2.cell(row=r, column=6, value=row["Narx"])
        ws2.cell(row=r, column=7, value=row["Summa"])
        ws2.cell(row=r, column=5).number_format = "#,##0.00"
        ws2.cell(row=r, column=6).number_format = "#,##0.00"
        ws2.cell(row=r, column=7).number_format = "#,##0.00"

    widths = [5, 28, 40, 20, 12, 14, 15]
    for i, w in enumerate(widths, start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    summary = {
        "total1": total1,
        "total2": total2,
        "diff": total2 - total1,
        "n_diff_rows": len(diff_df),
    }
    return buf.getvalue(), summary


# ================================================================
#            2) EXCEL FAYLNI TEKSHIRISH (validatsiya)
# ================================================================

def validate_file(pf: ParsedFile) -> tuple[bytes, int]:
    """qty * price != sum bo'lgan qatorlarni topadi (2 so'mgacha tolerantlik bilan)."""
    df = pf.df.copy()
    df["hisoblangan_summa"] = (df["qty"] * df["price"]).round(2)
    df["farq"] = (df["hisoblangan_summa"] - df["sum"]).round(2)
    bad = df[df["farq"].abs() > 2].copy()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Xatoliklar"
    bold = Font(name="Arial", bold=True)
    headers = ["№", "Nomi", "Kod", "Miqdor", "Narx", "Faylda summa", "Hisoblangan summa", "Farq"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = bold
        cell.fill = PatternFill("solid", fgColor="FCE4E4")

    for i, (_, row) in enumerate(bad.iterrows()):
        r = i + 2
        ws.cell(row=r, column=1, value=i + 1)
        ws.cell(row=r, column=2, value=row["name"])
        ws.cell(row=r, column=3, value=row["code"])
        ws.cell(row=r, column=4, value=row["qty"])
        ws.cell(row=r, column=5, value=row["price"])
        ws.cell(row=r, column=6, value=row["sum"])
        ws.cell(row=r, column=7, value=row["hisoblangan_summa"])
        ws.cell(row=r, column=8, value=row["farq"])

    widths = [5, 40, 16, 12, 14, 15, 18, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue(), len(bad)


# ================================================================
#   3) FAKTURALAR ASOSIDA CHORAK BO'YICHA FOYDA SOLIG'I HISOBOTI
# ================================================================

DOCS_COLUMN_MAP = {
    "seller_inn": "Продавец (ИНН или ПИНФЛ)",
    "seller_name": "Продавец (наименование)",
    "buyer_inn": "Покупатель (ИНН или ПИНФЛ)",
    "buyer_name": "Покупатель (наименование)",
    "doc_number": "Номер документ",
    "doc_date": "Дата документ",
    "description": "Примечание к товару (работе, услуге)",
    "summa_no_vat": "Стоимость поставки",
    "vat_rate": "НДС ставка",
    "vat_sum": "НДС сумма",
    "summa_with_vat": "Стоимость поставки с учётом НДС",
}


def _normalize_inn(x):
    """STIR/INN qiymatini tozalaydi: '303020732.0' -> '303020732', bo'sh -> ''."""
    if x is None or (isinstance(x, float) and pd.isna(x)):
        return ""
    s = str(x).strip()
    if s.lower() in ("nan", "none", ""):
        return ""
    if s.endswith(".0"):
        s = s[:-2]
    return re.sub(r"\D", "", s)


def parse_docs_export(path: str) -> pd.DataFrame:
    """soliqservis.uz / didox.uz dan yuklab olinadigan 'docs' eksport faylini o'qiydi."""
    df_raw = None
    for header_row in (1, 0, 2):
        try:
            candidate = pd.read_excel(path, header=header_row)
            if DOCS_COLUMN_MAP["seller_inn"] in candidate.columns:
                df_raw = candidate
                break
        except Exception:
            continue

    if df_raw is None:
        raise ValueError(
            "Fayl formati tanilmadi. Bu 'docs' eksport fayli (soliqservis.uz/didox.uz) "
            "ekanligiga ishonch hosil qiling."
        )

    missing = [v for v in DOCS_COLUMN_MAP.values() if v not in df_raw.columns]
    if missing:
        raise ValueError("Faylda kerakli ustunlar topilmadi: " + ", ".join(missing))

    out = pd.DataFrame()
    out["seller_inn"] = df_raw[DOCS_COLUMN_MAP["seller_inn"]].apply(_normalize_inn)
    out["seller_name"] = df_raw[DOCS_COLUMN_MAP["seller_name"]].astype(str).str.strip()
    out["buyer_inn"] = df_raw[DOCS_COLUMN_MAP["buyer_inn"]].apply(_normalize_inn)
    out["buyer_name"] = df_raw[DOCS_COLUMN_MAP["buyer_name"]].astype(str).str.strip()
    out["doc_number"] = df_raw[DOCS_COLUMN_MAP["doc_number"]].astype(str).str.strip()
    out["doc_date"] = pd.to_datetime(df_raw[DOCS_COLUMN_MAP["doc_date"]], format="%d-%m-%Y", errors="coerce")
    out["description"] = df_raw[DOCS_COLUMN_MAP["description"]].astype(str).str.strip()
    out["summa_no_vat"] = df_raw[DOCS_COLUMN_MAP["summa_no_vat"]].apply(_to_number).fillna(0.0)
    out["vat_rate"] = df_raw[DOCS_COLUMN_MAP["vat_rate"]].apply(_to_number).fillna(0.0)
    out["vat_sum"] = df_raw[DOCS_COLUMN_MAP["vat_sum"]].apply(_to_number).fillna(0.0)
    out["summa_with_vat"] = df_raw[DOCS_COLUMN_MAP["summa_with_vat"]].apply(_to_number).fillna(0.0)
    out = out.dropna(subset=["doc_date"])
    return out


QUARTER_MONTHS = {
    "1": (1, 3),
    "2": (4, 6),
    "3": (7, 9),
    "4": (10, 12),
}
QUARTER_LABELS = {
    "1": "I-chorak (yanvar-mart)",
    "2": "II-chorak (aprel-iyun)",
    "3": "III-chorak (iyul-sentabr)",
    "4": "IV-chorak (oktabr-dekabr)",
    "all": "Yil boshidan hozirgacha",
}


def compute_tax_report(answers: dict, df: pd.DataFrame):
    own_inn = re.sub(r"\D", "", answers["inn"])
    chorak = answers["chorak"]

    if df.empty:
        return "❌ Faylda sanasi to'g'ri o'qilgan qatorlar topilmadi."

    year = int(df["doc_date"].dt.year.max())

    if chorak == "all":
        start = datetime(year, 1, 1)
        end = df["doc_date"].max()
    else:
        m1, m2 = QUARTER_MONTHS[chorak]
        start = datetime(year, m1, 1)
        end_month_last_day = 31 if m2 in (1, 3, 5, 7, 8, 10, 12) else (30 if m2 != 2 else 28)
        end = datetime(year, m2, end_month_last_day, 23, 59, 59)

    period_df = df[(df["doc_date"] >= start) & (df["doc_date"] <= end)]

    income_rows = period_df[period_df["seller_inn"] == own_inn]
    expense_rows = period_df[period_df["buyer_inn"] == own_inn]
    matched_ids = set(income_rows.index) | set(expense_rows.index)
    unmatched = period_df[~period_df.index.isin(matched_ids)]

    daromad = income_rows["summa_no_vat"].sum()
    xarajat = expense_rows["summa_no_vat"].sum()
    qqs_daromad = income_rows["vat_sum"].sum()
    qqs_xarajat = expense_rows["vat_sum"].sum()
    baza = daromad - xarajat
    soliq = baza * 0.15 if baza > 0 else 0.0

    # ---- Excel hisobot ----
    wb = openpyxl.Workbook()
    bold = Font(name="Arial", bold=True)
    title_font = Font(name="Arial", bold=True, size=13)
    header_fill = PatternFill("solid", fgColor="D9E1F2")

    ws = wb.active
    ws.title = "Xulosa"
    ws["A1"] = f"Foyda solig'i hisoboti — {QUARTER_LABELS[chorak]} ({year})"
    ws["A1"].font = title_font
    ws.merge_cells("A1:B1")
    ws["A2"] = f"Kompaniya STIR: {own_inn}"

    rows = [
        ("Davr boshi", start.strftime("%d.%m.%Y")),
        ("Davr oxiri", end.strftime("%d.%m.%Y")),
        ("", ""),
        ("Jami daromad (QQS'siz)", round(daromad, 2)),
        ("Jami xarajat (QQS'siz)", round(xarajat, 2)),
        ("Soliq solinadigan foyda", round(baza, 2)),
        ("Foyda solig'i (15%)", round(soliq, 2)),
        ("", ""),
        ("Sotuvlar bo'yicha QQS", round(qqs_daromad, 2)),
        ("Xaridlar bo'yicha QQS", round(qqs_xarajat, 2)),
        ("", ""),
        ("Daromad fakturalari soni", len(income_rows)),
        ("Xarajat fakturalari soni", len(expense_rows)),
        ("Aniqlanmagan (na sotuvchi, na xaridor mos kelmadi)", len(unmatched)),
    ]
    r = 4
    for label, val in rows:
        ws.cell(row=r, column=1, value=label).font = Font(name="Arial", bold=label not in ("", None))
        ws.cell(row=r, column=2, value=val)
        r += 1
    ws.column_dimensions["A"].width = 48
    ws.column_dimensions["B"].width = 20

    def _write_invoice_sheet(name, rows_df):
        sh = wb.create_sheet(name)
        headers = ["Sana", "Hujjat №", "Kontragent", "Tavsif", "Summa (QQS'siz)", "QQS", "Jami (QQS bilan)"]
        sh.append(headers)
        for c in range(1, len(headers) + 1):
            cell = sh.cell(row=1, column=c)
            cell.font = bold
            cell.fill = header_fill
        for i, (_, row) in enumerate(rows_df.sort_values("doc_date").iterrows()):
            rr = i + 2
            kontragent = row["buyer_name"] if name == "Daromad fakturalari" else row["seller_name"]
            sh.cell(row=rr, column=1, value=row["doc_date"].strftime("%d.%m.%Y"))
            sh.cell(row=rr, column=2, value=row["doc_number"])
            sh.cell(row=rr, column=3, value=kontragent)
            sh.cell(row=rr, column=4, value=row["description"])
            sh.cell(row=rr, column=5, value=round(row["summa_no_vat"], 2))
            sh.cell(row=rr, column=6, value=round(row["vat_sum"], 2))
            sh.cell(row=rr, column=7, value=round(row["summa_with_vat"], 2))
        widths = [12, 26, 32, 40, 16, 14, 16]
        for i, w in enumerate(widths, start=1):
            sh.column_dimensions[get_column_letter(i)].width = w
        sh.freeze_panes = "A2"

    _write_invoice_sheet("Daromad fakturalari", income_rows)
    _write_invoice_sheet("Xarajat fakturalari", expense_rows)
    if len(unmatched) > 0:
        _write_invoice_sheet("Aniqlanmagan", unmatched)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    text = (
        f"📑 Foyda solig'i hisoboti — {QUARTER_LABELS[chorak]} ({year})\n\n"
        f"Davr: {start.strftime('%d.%m.%Y')} — {end.strftime('%d.%m.%Y')}\n\n"
        f"Jami daromad (QQS'siz): {daromad:,.0f} so'm ({len(income_rows)} ta faktura)\n"
        f"Jami xarajat (QQS'siz): {xarajat:,.0f} so'm ({len(expense_rows)} ta faktura)\n"
        f"Soliq solinadigan foyda: {baza:,.0f} so'm\n"
        f"✅ Foyda solig'i (15%): {soliq:,.0f} so'm\n"
    )
    if len(unmatched) > 0:
        text += (
            f"\n⚠️ {len(unmatched)} ta qatorda STIR na sotuvchi, na xaridur sifatida topilmadi "
            f"— ular alohida 'Aniqlanmagan' varag'ida ko'rsatildi, hisobga kiritilmadi."
        )
    text += "\n\nTo'liq hisobotni Excel faylida yuboryapman 👇"

    return text, buf.getvalue(), f"foyda_soligi_{QUARTER_LABELS[chorak].split()[0]}_{year}.xlsx"


# ================================================================
#        4) GENERIK QADAM-BA-QADAM HISOB-KITOB MOTORI
# ================================================================

def parse_value(raw: str, vtype: str):
    """(qiymat, xato_matni) qaytaradi. Muvaffaqiyatli bo'lsa xato_matni None."""
    raw = raw.strip()
    if vtype == "float":
        s = raw.replace(" ", "").replace(",", ".")
        s = re.sub(r"[^0-9.\-]", "", s)
        try:
            return float(s), None
        except ValueError:
            return None, "❌ Iltimos, faqat raqam kiriting (masalan: 1500000). Qayta urinib ko'ring:"
    if vtype == "int":
        s = re.sub(r"[^0-9\-]", "", raw)
        try:
            return int(s), None
        except ValueError:
            return None, "❌ Iltimos, faqat butun son kiriting (masalan: 12). Qayta urinib ko'ring:"
    if vtype == "date":
        try:
            return datetime.strptime(raw, "%d.%m.%Y"), None
        except ValueError:
            return None, "❌ Sana formati noto'g'ri. Masalan: 01.07.2026. Qayta urinib ko'ring:"
    if vtype == "text":
        if not raw:
            return None, "❌ Matn bo'sh bo'lmasligi kerak. Qayta kiriting:"
        return raw, None
    if vtype == "invoice_rows":
        lines = [l.strip() for l in raw.split("\n") if l.strip()]
        rows = []
        for line in lines:
            parts = [p.strip() for p in line.split(";")]
            if len(parts) != 3:
                continue
            name, qty_s, price_s = parts
            try:
                qty = float(qty_s.replace(",", "."))
                price = float(price_s.replace(",", "."))
            except ValueError:
                continue
            rows.append((name, qty, price))
        if not rows:
            return None, (
                "❌ Hech qanday to'g'ri qator topilmadi.\n"
                "Format: Nomi;Miqdor;Narx (har biri alohida qatorda). Qayta yuboring:"
            )
        return rows, None
    return raw, None


# ---------------- Hisob-kitob funksiyalari ----------------

def compute_profit_tax(a):
    daromad, xarajat = a["daromad"], a["xarajat"]
    baza = daromad - xarajat
    if baza <= 0:
        return (
            f"📊 Natija:\n\nDaromad: {daromad:,.0f} so'm\nXarajat: {xarajat:,.0f} so'm\n\n"
            f"Soliq solinadigan foyda mavjud emas (zarar yoki nol). Foyda solig'i: 0 so'm."
        )
    soliq = baza * 0.15
    return (
        f"💰 Foyda solig'i hisob-kitobi\n\n"
        f"Jami daromad: {daromad:,.0f} so'm\n"
        f"Xarajatlar: {xarajat:,.0f} so'm\n"
        f"Soliq solinadigan foyda: {baza:,.0f} so'm\n"
        f"Stavka: 15% (asosiy)\n\n"
        f"✅ Foyda solig'i: {soliq:,.0f} so'm\n\n"
        f"ℹ️ Eslatma: bir qator sohalar uchun imtiyozli stavkalar (masalan 2%, 0%) amal qiladi. "
        f"Aniq stavkani soliq.uz orqali tekshiring."
    )


def compute_vat(a):
    yonalish, summa = a["yonalish"], a["summa"]
    rate = 0.12
    if yonalish == "add":
        qqs = summa * rate
        jami = summa + qqs
        return (
            f"🧮 QQS kalkulyatori\n\n"
            f"QQS'siz summa: {summa:,.0f} so'm\n"
            f"QQS (12%): {qqs:,.0f} so'm\n"
            f"✅ QQS bilan jami: {jami:,.0f} so'm"
        )
    else:
        asosiy = summa / 1.12
        qqs = summa - asosiy
        return (
            f"🧮 QQS kalkulyatori\n\n"
            f"QQS bilan summa: {summa:,.0f} so'm\n"
            f"Shundan QQS (12%): {qqs:,.0f} so'm\n"
            f"✅ QQS'siz summa: {asosiy:,.0f} so'm"
        )


def compute_salary(a):
    yalpi = a["yalpi"]
    ndfl = yalpi * 0.12
    qolga = yalpi - ndfl
    return (
        f"💵 Ish haqi kalkulyatori\n\n"
        f"Yalpi ish haqi: {yalpi:,.0f} so'm\n"
        f"Jismoniy shaxslardan daromad solig'i (12%): {ndfl:,.0f} so'm\n"
        f"✅ Qo'lga tegadigan summa: {qolga:,.0f} so'm\n\n"
        f"ℹ️ Eslatma: ijtimoiy soliq (12%) ish beruvchi tomonidan alohida to'lanadi, "
        f"xodim ish haqidan ushlanmaydi. Chegirmalar (BHM va h.k.) hisobga olinmagan."
    )


def compute_penalty(a):
    qarz, kunlar, stavka = a["qarz"], a["kunlar"], a["stavka"]
    penya = qarz * (stavka / 100) * kunlar
    return (
        f"📉 Penya (jarima) kalkulyatori\n\n"
        f"Qarz summasi: {qarz:,.0f} so'm\n"
        f"Kechiktirilgan kunlar: {kunlar}\n"
        f"Kunlik stavka: {stavka}%\n\n"
        f"✅ Jami penya: {penya:,.0f} so'm"
    )


def compute_depreciation(a):
    narx, muddat = a["narx"], a["muddat"]
    if muddat <= 0:
        return "❌ Foydali muddat noldan katta bo'lishi kerak."
    yillik = narx / muddat
    oylik = yillik / 12
    return (
        f"📉 Amortizatsiya kalkulyatori (chiziqli usul)\n\n"
        f"Boshlang'ich qiymat: {narx:,.0f} so'm\n"
        f"Foydali muddat: {muddat:g} yil\n\n"
        f"✅ Yillik amortizatsiya: {yillik:,.0f} so'm\n"
        f"✅ Oylik amortizatsiya: {oylik:,.0f} so'm"
    )


def compute_currency(a):
    code = a["valyuta"]
    try:
        resp = requests.get("https://cbu.uz/uz/arkhiv-kursov-valyut/json/", timeout=10)
        data = resp.json()
        item = next((d for d in data if d.get("Ccy") == code), None)
        if not item:
            return f"❌ {code} valyutasi bo'yicha ma'lumot topilmadi."
        return (
            f"💱 CBU rasmiy kursi ({item.get('Date')})\n\n"
            f"1 {code} = {item.get('Rate')} so'm\n"
            f"Kunlik o'zgarish: {item.get('Diff')} so'm"
        )
    except Exception as e:
        logger.exception("CBU kursini olishda xatolik")
        return f"❌ Kursni olishda xatolik yuz berdi: {e}\nBirozdan so'ng qayta urinib ko'ring."


def compute_workdays(a):
    start, end = a["start"], a["end"]
    if end < start:
        start, end = end, start
    jami_kun = (end - start).days + 1
    dam_olish = 0
    ish_kun = 0
    cur = start
    from datetime import timedelta
    for _ in range(jami_kun):
        if cur.weekday() >= 5:
            dam_olish += 1
        else:
            ish_kun += 1
        cur += timedelta(days=1)
    return (
        f"📆 Ish kunlari kalkulyatori\n\n"
        f"Davr: {start.strftime('%d.%m.%Y')} — {end.strftime('%d.%m.%Y')}\n"
        f"Jami kunlar: {jami_kun}\n"
        f"Dam olish kunlari (shanba/yakshanba): {dam_olish}\n"
        f"✅ Ish kunlari: {ish_kun}\n\n"
        f"ℹ️ Rasmiy bayram kunlari hisobga olinmagan."
    )


def compute_loan(a):
    summa, muddat, stavka = a["summa"], a["muddat"], a["stavka"]
    if muddat <= 0 or summa <= 0:
        return "❌ Summa va muddat noldan katta bo'lishi kerak."
    oylik_stavka = stavka / 100 / 12
    if oylik_stavka == 0:
        oylik_tolov = summa / muddat
    else:
        oylik_tolov = summa * oylik_stavka * (1 + oylik_stavka) ** muddat / ((1 + oylik_stavka) ** muddat - 1)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "To'lov jadvali"
    bold = Font(name="Arial", bold=True)
    headers = ["Oy", "Oy boshiga qoldiq", "Oylik to'lov", "Foiz qismi", "Asosiy qarz qismi", "Oy oxiriga qoldiq"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = bold
        cell.fill = PatternFill("solid", fgColor="D9E1F2")

    qoldiq = summa
    jami_foiz = 0.0
    for oy in range(1, int(muddat) + 1):
        foiz_qismi = qoldiq * oylik_stavka
        asosiy_qismi = oylik_tolov - foiz_qismi
        yangi_qoldiq = max(qoldiq - asosiy_qismi, 0)
        r = oy + 1
        ws.cell(row=r, column=1, value=oy)
        ws.cell(row=r, column=2, value=round(qoldiq, 2))
        ws.cell(row=r, column=3, value=round(oylik_tolov, 2))
        ws.cell(row=r, column=4, value=round(foiz_qismi, 2))
        ws.cell(row=r, column=5, value=round(asosiy_qismi, 2))
        ws.cell(row=r, column=6, value=round(yangi_qoldiq, 2))
        jami_foiz += foiz_qismi
        qoldiq = yangi_qoldiq

    widths = [6, 18, 16, 14, 18, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    text = (
        f"🏦 Kredit/lizing to'lov jadvali\n\n"
        f"Kredit summasi: {summa:,.0f} so'm\n"
        f"Muddat: {muddat} oy\n"
        f"Yillik stavka: {stavka}%\n\n"
        f"✅ Oylik to'lov (annuitet): {oylik_tolov:,.0f} so'm\n"
        f"Jami to'lanadigan foiz: {jami_foiz:,.0f} so'm\n\n"
        f"To'liq jadvalni Excel faylida yubordim 👇"
    )
    return text, buf.getvalue(), "kredit_jadvali.xlsx"


def compute_ratios(a):
    ja, jm, sf, ua = a["joriy_aktiv"], a["joriy_majburiyat"], a["sof_foyda"], a["umumiy_aktiv"]
    likvidlik = ja / jm if jm else None
    roa = (sf / ua * 100) if ua else None
    lines = ["📈 Moliyaviy koeffitsientlar\n"]
    if likvidlik is not None:
        holat = "yaxshi ✅" if likvidlik >= 1 else "e'tibor talab qiladi ⚠️"
        lines.append(f"Joriy likvidlik koeffitsienti: {likvidlik:.2f} ({holat})")
    if roa is not None:
        lines.append(f"Aktivlar rentabelligi (ROA): {roa:.2f}%")
    lines.append("\nℹ️ Bu — umumiy taxminiy ko'rsatkichlar, soha o'rtacha ko'rsatkichlari bilan solishtirib baholang.")
    return "\n".join(lines)


def compute_invoice(a):
    kompaniya = a["kompaniya"]
    rows = a["rows"]
    rate = 0.12

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hisob-faktura"
    bold = Font(name="Arial", bold=True)
    title_font = Font(name="Arial", bold=True, size=13)

    ws["A1"] = f"Hisob-faktura — {kompaniya}"
    ws["A1"].font = title_font
    ws.merge_cells("A1:F1")
    ws["A2"] = f"Sana: {datetime.now().strftime('%d.%m.%Y')}"

    headers = ["№", "Nomi", "Miqdor", "Narx (QQS'siz)", "Summa (QQS'siz)", "QQS (12%)", "Jami (QQS bilan)"]
    ws.append([])
    ws.append(headers)
    header_row = 4
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=c)
        cell.font = bold
        cell.fill = PatternFill("solid", fgColor="D9E1F2")

    jami_summa = 0.0
    jami_qqs = 0.0
    for i, (name, qty, price) in enumerate(rows):
        r = header_row + 1 + i
        summa = qty * price
        qqs = summa * rate
        jami = summa + qqs
        ws.cell(row=r, column=1, value=i + 1)
        ws.cell(row=r, column=2, value=name)
        ws.cell(row=r, column=3, value=qty)
        ws.cell(row=r, column=4, value=price)
        ws.cell(row=r, column=5, value=round(summa, 2))
        ws.cell(row=r, column=6, value=round(qqs, 2))
        ws.cell(row=r, column=7, value=round(jami, 2))
        jami_summa += summa
        jami_qqs += qqs

    total_row = header_row + 1 + len(rows)
    ws.cell(row=total_row, column=2, value="JAMI").font = bold
    ws.cell(row=total_row, column=5, value=round(jami_summa, 2)).font = bold
    ws.cell(row=total_row, column=6, value=round(jami_qqs, 2)).font = bold
    ws.cell(row=total_row, column=7, value=round(jami_summa + jami_qqs, 2)).font = bold

    widths = [5, 32, 10, 16, 16, 14, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    text = (
        f"📄 Hisob-faktura tayyor — {kompaniya}\n\n"
        f"Qatorlar soni: {len(rows)}\n"
        f"Jami summa (QQS'siz): {jami_summa:,.0f} so'm\n"
        f"QQS (12%): {jami_qqs:,.0f} so'm\n"
        f"✅ Jami (QQS bilan): {jami_summa + jami_qqs:,.0f} so'm"
    )
    return text, buf.getvalue(), "hisob-faktura.xlsx"


YATT_INFO_TEXT = (
    "🧾 YATT (Yakka tartibdagi tadbirkor) uchun soliq\n\n"
    "YATT uchun soliq odatda foizli emas, balki hudud va faoliyat turiga qarab "
    "belgilangan qat'iy (fiksirlangan) summa asosida hisoblanadi va vaqti-vaqti bilan o'zgaradi.\n\n"
    "Shu sababli bu yerda avtomatik hisoblash noaniq natija berishi mumkin. "
    "Aniq summani bilish uchun:\n"
    "• soliq.uz saytidagi rasmiy kalkulyatordan foydalaning, yoki\n"
    "• hududingizdagi Davlat soliq inspeksiyasiga murojaat qiling.\n\n"
    "Agar xohlasangiz, keyinchalik hududlar bo'yicha jadvalni botga qo'shib, "
    "aniqroq hisoblashni qo'shishimiz mumkin."
)


# ---------------- Bo'lim (flow) ta'riflari ----------------

FLOWS = {
    "profit_tax": {
        "title": "💰 Foyda solig'i hisob-kitobi",
        "steps": [
            {"key": "daromad", "prompt": "1/2. Jami daromad summasini kiriting (so'm):", "type": "float"},
            {"key": "xarajat", "prompt": "2/2. Jami (chegiriladigan) xarajatlar summasini kiriting (so'm):", "type": "float"},
        ],
        "compute": compute_profit_tax,
    },
    "vat": {
        "title": "🧮 QQS kalkulyatori",
        "steps": [
            {
                "key": "yonalish",
                "prompt": "Nima qilish kerak?",
                "type": "choice",
                "options": [
                    ("QQS'siz summaga QQS qo'shish", "add"),
                    ("QQS bilan summadan QQS ajratib olish", "extract"),
                ],
            },
            {"key": "summa", "prompt": "Summani kiriting (so'm):", "type": "float"},
        ],
        "compute": compute_vat,
    },
    "salary": {
        "title": "💵 Ish haqi kalkulyatori",
        "steps": [
            {"key": "yalpi", "prompt": "Yalpi (hisoblangan) ish haqini kiriting (so'm):", "type": "float"},
        ],
        "compute": compute_salary,
    },
    "penalty": {
        "title": "📉 Penya (jarima) kalkulyatori",
        "steps": [
            {"key": "qarz", "prompt": "1/3. Qarz (asosiy) summasini kiriting (so'm):", "type": "float"},
            {"key": "kunlar", "prompt": "2/3. Kechiktirilgan kunlar sonini kiriting:", "type": "int"},
            {"key": "stavka", "prompt": "3/3. Kunlik penya stavkasini foizda kiriting (masalan 0.045):", "type": "float"},
        ],
        "compute": compute_penalty,
    },
    "depreciation": {
        "title": "📉 Amortizatsiya kalkulyatori",
        "steps": [
            {"key": "narx", "prompt": "1/2. Asosiy vositaning boshlang'ich qiymatini kiriting (so'm):", "type": "float"},
            {"key": "muddat", "prompt": "2/2. Foydali xizmat muddatini (yillarda) kiriting:", "type": "float"},
        ],
        "compute": compute_depreciation,
    },
    "currency": {
        "title": "💱 CBU rasmiy valyuta kursi",
        "steps": [
            {
                "key": "valyuta",
                "prompt": "Qaysi valyuta kursini bilmoqchisiz?",
                "type": "choice",
                "options": [("USD", "USD"), ("EUR", "EUR"), ("RUB", "RUB"), ("GBP", "GBP")],
            },
        ],
        "compute": compute_currency,
    },
    "invoice": {
        "title": "📄 Hisob-faktura generator",
        "steps": [
            {"key": "kompaniya", "prompt": "1/2. Kompaniya nomini kiriting:", "type": "text"},
            {
                "key": "rows",
                "prompt": (
                    "2/2. Mahsulotlarni quyidagi formatda yuboring "
                    "(har biri alohida qatorda):\n\nNomi;Miqdor;Narx\n\n"
                    "Masalan:\nNoutbuk;2;8500000\nSichqoncha;5;120000"
                ),
                "type": "invoice_rows",
            },
        ],
        "compute": compute_invoice,
    },
    "workdays": {
        "title": "📆 Ish kunlari kalkulyatori",
        "steps": [
            {"key": "start", "prompt": "1/2. Boshlanish sanasini kiriting (KK.OO.YYYY, masalan 01.07.2026):", "type": "date"},
            {"key": "end", "prompt": "2/2. Tugash sanasini kiriting (KK.OO.YYYY):", "type": "date"},
        ],
        "compute": compute_workdays,
    },
    "loan": {
        "title": "🏦 Kredit/lizing to'lov jadvali",
        "steps": [
            {"key": "summa", "prompt": "1/3. Kredit summasini kiriting (so'm):", "type": "float"},
            {"key": "muddat", "prompt": "2/3. Kredit muddatini (oylarda) kiriting:", "type": "int"},
            {"key": "stavka", "prompt": "3/3. Yillik foiz stavkasini kiriting (%):", "type": "float"},
        ],
        "compute": compute_loan,
    },
    "ratios": {
        "title": "📈 Moliyaviy koeffitsientlar",
        "steps": [
            {"key": "joriy_aktiv", "prompt": "1/4. Joriy aktivlar summasini kiriting:", "type": "float"},
            {"key": "joriy_majburiyat", "prompt": "2/4. Joriy majburiyatlar summasini kiriting:", "type": "float"},
            {"key": "sof_foyda", "prompt": "3/4. Sof foyda summasini kiriting:", "type": "float"},
            {"key": "umumiy_aktiv", "prompt": "4/4. Umumiy aktivlar summasini kiriting:", "type": "float"},
        ],
        "compute": compute_ratios,
    },
    "tax_report": {
        "title": "📑 Fakturalar asosida foyda solig'i",
        "steps": [
            {"key": "inn", "prompt": "1/2. Kompaniyangizning STIR (INN) raqamini kiriting (9 xonali):", "type": "text"},
            {
                "key": "chorak",
                "prompt": "2/2. Qaysi davr uchun hisoblaymiz?",
                "type": "choice",
                "options": [
                    ("I-chorak (yanvar-mart)", "1"),
                    ("II-chorak (aprel-iyun)", "2"),
                    ("III-chorak (iyul-sentabr)", "3"),
                    ("IV-chorak (oktabr-dekabr)", "4"),
                    ("Yil boshidan hozirgacha", "all"),
                ],
            },
        ],
        "compute": None,
        "needs_file": True,
    },
}


MAIN_MENU_LAYOUT = [
    [("📊 Fayllarni solishtirish", "menu:compare"), ("🔍 Faylni tekshirish", "menu:validate")],
    [("💰 Foyda solig'i", "menu:profit_tax"), ("🧮 QQS kalkulyatori", "menu:vat")],
    [("💵 Ish haqi kalkulyatori", "menu:salary"), ("📉 Penya kalkulyatori", "menu:penalty")],
    [("📉 Amortizatsiya", "menu:depreciation"), ("💱 CBU valyuta kursi", "menu:currency")],
    [("📄 Hisob-faktura", "menu:invoice"), ("📆 Ish kunlari", "menu:workdays")],
    [("🏦 Kredit jadvali", "menu:loan"), ("📈 Moliyaviy koeffitsientlar", "menu:ratios")],
    [("🧾 YATT uchun soliq", "menu:yatt")],
    [("📑 Fakturalar → foyda solig'i (chorak)", "menu:tax_report")],
]


def _menu_markup():
    kb = [[InlineKeyboardButton(label, callback_data=cd) for label, cd in row] for row in MAIN_MENU_LAYOUT]
    return InlineKeyboardMarkup(kb)


async def show_main_menu(chat_target, text):
    """chat_target — update.message yoki callback_query.message bo'lishi mumkin (reply_text bor)."""
    await chat_target.reply_text(text, reply_markup=_menu_markup())


# ================================================================
#                    TELEGRAM HANDLERLAR
# ================================================================

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data.clear()
    await update.message.reply_text(
        "Salom! 👋 Men buxgalteriya va biznes hisob-kitoblari uchun yordamchi botman.\n\n"
        "Kerakli bo'limni tanlang:",
        reply_markup=_menu_markup(),
    )
    return MENU


async def ask_current_step(query, context):
    flow = FLOWS[context.user_data["flow"]]
    idx = context.user_data["step_index"]
    step = flow["steps"][idx]
    text = step["prompt"]
    if step["type"] == "choice":
        kb = [[InlineKeyboardButton(label, callback_data=f"choice:{val}")] for label, val in step["options"]]
        await query.message.reply_text(text, reply_markup=InlineKeyboardMarkup(kb))
    else:
        await query.message.reply_text(text)


async def menu_router(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    choice = query.data.split(":", 1)[1]

    if choice == "compare":
        await query.message.reply_text("Birinchi faylni yuboring (.xlsx yoki .xls).")
        return WAITING_FILE_1

    if choice == "validate":
        await query.message.reply_text("Tekshiriladigan Excel faylni yuboring (.xlsx yoki .xls).")
        return WAITING_VALIDATE_FILE

    if choice == "yatt":
        await query.message.reply_text(YATT_INFO_TEXT)
        await show_main_menu(query.message, "Boshqa bo'limni tanlashingiz mumkin:")
        return MENU

    if choice in FLOWS:
        context.user_data["flow"] = choice
        context.user_data["step_index"] = 0
        context.user_data["answers"] = {}
        await ask_current_step(query, context)
        return TEXT_INPUT

    await show_main_menu(query.message, "Bo'lim tanlang:")
    return MENU


async def _finish_flow(reply_target, context):
    """reply_target — reply_text/reply_document metodlariga ega obyekt (Message)."""
    flow = FLOWS[context.user_data["flow"]]
    result = flow["compute"](context.user_data["answers"])
    if isinstance(result, tuple):
        text, file_bytes, filename = result
        await reply_target.reply_text(text)
        await reply_target.reply_document(document=io.BytesIO(file_bytes), filename=filename)
    else:
        await reply_target.reply_text(result)
    context.user_data.clear()
    await show_main_menu(reply_target, "Yana biror bo'limni tanlashingiz mumkin:")


async def handle_choice(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if "flow" not in context.user_data:
        await query.message.reply_text("Iltimos, avval /start orqali menyuni oching.")
        return MENU
    val = query.data.split(":", 1)[1]
    flow = FLOWS[context.user_data["flow"]]
    idx = context.user_data["step_index"]
    step = flow["steps"][idx]
    context.user_data["answers"][step["key"]] = val
    context.user_data["step_index"] += 1

    if context.user_data["step_index"] < len(flow["steps"]):
        await ask_current_step(query, context)
        return TEXT_INPUT
    else:
        return await _steps_complete(query.message, context, flow)


async def _steps_complete(message, context, flow):
    """Barcha qadamlar to'ldirilgach chaqiriladi: yoki hisoblaydi, yoki fayl so'raydi."""
    if flow.get("needs_file"):
        await message.reply_text("Endi kerakli Excel faylni yuboring (.xlsx).")
        return WAITING_TAX_REPORT_FILE
    await _finish_flow(message, context)
    return MENU


async def handle_text_input(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if "flow" not in context.user_data:
        await update.message.reply_text("Iltimos, avval /start orqali menyuni oching.")
        return MENU

    flow = FLOWS[context.user_data["flow"]]
    idx = context.user_data["step_index"]
    step = flow["steps"][idx]
    raw = update.message.text
    value, error = parse_value(raw, step["type"])
    if error:
        await update.message.reply_text(error)
        return TEXT_INPUT

    context.user_data["answers"][step["key"]] = value
    context.user_data["step_index"] += 1

    if context.user_data["step_index"] < len(flow["steps"]):
        await ask_current_step_message(update.message, context)
        return TEXT_INPUT
    else:
        return await _steps_complete(update.message, context, flow)


async def ask_current_step_message(message, context):
    flow = FLOWS[context.user_data["flow"]]
    idx = context.user_data["step_index"]
    step = flow["steps"][idx]
    text = step["prompt"]
    if step["type"] == "choice":
        kb = [[InlineKeyboardButton(label, callback_data=f"choice:{val}")] for label, val in step["options"]]
        await message.reply_text(text, reply_markup=InlineKeyboardMarkup(kb))
    else:
        await message.reply_text(text)


# ---------------- Fayllarni solishtirish handlerlari ----------------

async def receive_file_1(update: Update, context: ContextTypes.DEFAULT_TYPE):
    doc: Document = update.message.document
    if not doc or not doc.file_name.lower().endswith((".xlsx", ".xls", ".xlsm")):
        await update.message.reply_text("Iltimos, .xlsx yoki .xls fayl yuboring.")
        return WAITING_FILE_1

    path = f"/tmp/{update.effective_chat.id}_1_{doc.file_name}"
    file = await doc.get_file()
    await file.download_to_drive(path)
    context.user_data["file1_path"] = path

    await update.message.reply_text("1-fayl qabul qilindi ✅\nEndi ikkinchi faylni yuboring.")
    return WAITING_FILE_2


async def receive_file_2(update: Update, context: ContextTypes.DEFAULT_TYPE):
    doc: Document = update.message.document
    if not doc or not doc.file_name.lower().endswith((".xlsx", ".xls", ".xlsm")):
        await update.message.reply_text("Iltimos, .xlsx yoki .xls fayl yuboring.")
        return WAITING_FILE_2

    path = f"/tmp/{update.effective_chat.id}_2_{doc.file_name}"
    file = await doc.get_file()
    await file.download_to_drive(path)

    await update.message.reply_text("2-fayl qabul qilindi ✅. Solishtirilmoqda...")

    try:
        pf1 = read_any(context.user_data["file1_path"])
        pf2 = read_any(path)
        report_bytes, summary = compare_files(pf1, pf2)
    except Exception as e:
        logger.exception("Solishtirishda xatolik")
        await update.message.reply_text(f"❌ Xatolik yuz berdi: {e}")
        context.user_data.clear()
        await show_main_menu(update.message, "Menyuga qaytdik:")
        return MENU

    text = (
        f"✅ Solishtirish tugadi.\n\n"
        f"1-fayl jami: {summary['total1']:,.2f}\n"
        f"2-fayl jami: {summary['total2']:,.2f}\n"
        f"Umumiy farq: {summary['diff']:,.2f}\n"
        f"Farqli qatorlar soni: {summary['n_diff_rows']}\n\n"
        f"To'liq hisobot faylini yuboryapman 👇"
    )
    await update.message.reply_text(text)
    await update.message.reply_document(
        document=io.BytesIO(report_bytes), filename="farqlar_hisoboti.xlsx"
    )

    context.user_data.clear()
    await show_main_menu(update.message, "Yana biror bo'limni tanlashingiz mumkin:")
    return MENU


async def receive_validate_file(update: Update, context: ContextTypes.DEFAULT_TYPE):
    doc: Document = update.message.document
    if not doc or not doc.file_name.lower().endswith((".xlsx", ".xls", ".xlsm")):
        await update.message.reply_text("Iltimos, .xlsx yoki .xls fayl yuboring.")
        return WAITING_VALIDATE_FILE

    path = f"/tmp/{update.effective_chat.id}_v_{doc.file_name}"
    file = await doc.get_file()
    await file.download_to_drive(path)

    await update.message.reply_text("Fayl qabul qilindi ✅. Tekshirilmoqda...")

    try:
        pf = read_any(path)
        report_bytes, n_bad = validate_file(pf)
    except Exception as e:
        logger.exception("Tekshirishda xatolik")
        await update.message.reply_text(f"❌ Xatolik yuz berdi: {e}")
        context.user_data.clear()
        await show_main_menu(update.message, "Menyuga qaytdik:")
        return MENU

    if n_bad == 0:
        await update.message.reply_text("✅ Tekshirildi. Arifmetik xatolik topilmadi — hammasi joyida!")
    else:
        await update.message.reply_text(
            f"⚠️ Tekshirildi. {n_bad} ta qatorda miqdor × narx ≠ summa nomuvofiqligi topildi.\n"
            f"To'liq ro'yxatni Excel faylida yuboryapman 👇"
        )
        await update.message.reply_document(
            document=io.BytesIO(report_bytes), filename="tekshiruv_natijasi.xlsx"
        )

    context.user_data.clear()
    await show_main_menu(update.message, "Yana biror bo'limni tanlashingiz mumkin:")
    return MENU


async def receive_tax_report_file(update: Update, context: ContextTypes.DEFAULT_TYPE):
    doc: Document = update.message.document
    if not doc or not doc.file_name.lower().endswith((".xlsx", ".xls", ".xlsm")):
        await update.message.reply_text("Iltimos, .xlsx fayl yuboring.")
        return WAITING_TAX_REPORT_FILE

    path = f"/tmp/{update.effective_chat.id}_tr_{doc.file_name}"
    file = await doc.get_file()
    await file.download_to_drive(path)

    await update.message.reply_text("Fayl qabul qilindi ✅. Hisoblanmoqda...")

    try:
        df = parse_docs_export(path)
        result = compute_tax_report(context.user_data.get("answers", {}), df)
    except Exception as e:
        logger.exception("Foyda solig'i hisobotida xatolik")
        await update.message.reply_text(f"❌ Xatolik yuz berdi: {e}")
        context.user_data.clear()
        await show_main_menu(update.message, "Menyuga qaytdik:")
        return MENU

    if isinstance(result, tuple):
        text, file_bytes, filename = result
        await update.message.reply_text(text)
        await update.message.reply_document(document=io.BytesIO(file_bytes), filename=filename)
    else:
        await update.message.reply_text(result)

    context.user_data.clear()
    await show_main_menu(update.message, "Yana biror bo'limni tanlashingiz mumkin:")
    return MENU


async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data.clear()
    await update.message.reply_text("Bekor qilindi.")
    await show_main_menu(update.message, "Kerakli bo'limni tanlang:")
    return MENU


def main():
    if not BOT_TOKEN:
        raise RuntimeError("BOT_TOKEN muhit o'zgaruvchisi topilmadi. `export BOT_TOKEN=...` qiling.")

    app = Application.builder().token(BOT_TOKEN).build()

    conv = ConversationHandler(
        entry_points=[CommandHandler("start", start)],
        states={
            MENU: [CallbackQueryHandler(menu_router, pattern="^menu:")],
            TEXT_INPUT: [
                CallbackQueryHandler(handle_choice, pattern="^choice:"),
                MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text_input),
            ],
            WAITING_FILE_1: [MessageHandler(filters.Document.ALL, receive_file_1)],
            WAITING_FILE_2: [MessageHandler(filters.Document.ALL, receive_file_2)],
            WAITING_VALIDATE_FILE: [MessageHandler(filters.Document.ALL, receive_validate_file)],
            WAITING_TAX_REPORT_FILE: [MessageHandler(filters.Document.ALL, receive_tax_report_file)],
        },
        fallbacks=[CommandHandler("cancel", cancel), CommandHandler("start", start)],
    )
    app.add_handler(conv)

    logger.info("Bot ishga tushdi...")
    app.run_polling()


if __name__ == "__main__":
    main()
